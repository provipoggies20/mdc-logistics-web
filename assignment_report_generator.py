#!/usr/bin/env python3
import mysql.connector
from mysql.connector import Error
import pandas as pd
import math
from datetime import datetime
import logging
import os
import sys
import json

# Setup logging
current_directory = os.path.dirname(os.path.abspath(__file__))
logs_directory = os.path.join(current_directory, 'Logs')
os.makedirs(logs_directory, exist_ok=True)
log_file_path = os.path.join(logs_directory, 'vehicle_info_report.log')

logging.basicConfig(
    filename=log_file_path,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

def log_message(message):
    """Log message only (no print to avoid corrupting output)"""
    logging.info(message)

def create_connection():
    """Create database connection"""
    try:
        connection = mysql.connector.connect(
            host='localhost',
            database='mdc',
            user='root',
            password='',
            connection_timeout=300
        )
        if connection.is_connected():
            log_message("Connected to MySQL database")
            return connection
    except Error as e:
        log_message(f"Error connecting to database: {e}")
        return None

def haversine(coord1, coord2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    Returns distance in kilometers
    """
    R = 6371.0  # Radius of the Earth in kilometers
    lat1, lon1 = coord1
    lat2, lon2 = coord2
    
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)
    
    dlon = lon2_rad - lon1_rad
    dlat = lat2_rad - lat1_rad
    
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.asin(math.sqrt(a))
    
    return R * c

def is_within_geofence(device_location, geofence_coordinates, radius=5):
    """Check if device is within geofence radius (default 5km)"""
    distance = haversine(device_location, geofence_coordinates)
    return distance <= radius

def get_assignment_tables(connection):
    """Get all tables that start with 'assignment_'"""
    try:
        cursor = connection.cursor()
        query = """
        SELECT table_name 
        FROM information_schema.tables 
        WHERE table_schema = 'mdc'
        AND table_name LIKE 'assignment_%'
        """
        cursor.execute(query)
        tables = [table[0] for table in cursor.fetchall()]
        log_message(f"Found {len(tables)} assignment tables")
        return tables
    except Error as e:
        log_message(f"Error getting assignment tables: {e}")
        return []

def get_assignment_coordinates(connection, table_name):
    """Get coordinates and site names from assignment table"""
    try:
        cursor = connection.cursor()
        query = f"SELECT site, coordinates, location FROM {table_name}"
        cursor.execute(query)
        results = cursor.fetchall()
        
        coordinates_data = []
        for row in results:
            site = row[0]
            coords_str = row[1].strip()
            location = row[2] if row[2] else ""
            
            try:
                # Parse coordinates (format: "lat, lon" or "lat,lon")
                coords_parts = coords_str.replace('\n', '').split(',')
                if len(coords_parts) == 2:
                    lat = float(coords_parts[0].strip())
                    lon = float(coords_parts[1].strip())
                    coordinates_data.append({
                        'site': site,
                        'lat': lat,
                        'lon': lon,
                        'location': location
                    })
            except (ValueError, IndexError) as e:
                log_message(f"Invalid coordinate format in {table_name} for site {site}: {coords_str}")
                continue
        
        return coordinates_data
    except Error as e:
        log_message(f"Error getting coordinates from {table_name}: {e}")
        return []

def get_devices_data(connection):
    """Get all devices from devices table with necessary columns including tag, status, remarks, days_no_gps, and last_gps_assignment"""
    try:
        cursor = connection.cursor(dictionary=True)
        query = """
        SELECT 
            target_name,
            latitude,
            longitude,
            address,
            cut_address,
            equipment_type,
            assignment as current_assignment,
            tag,
            physical_status,
            remarks,
            days_no_gps,
            last_gps_assignment
        FROM devices
        WHERE latitude IS NOT NULL 
        AND longitude IS NOT NULL
        AND latitude != 0 
        AND longitude != 0
        """
        cursor.execute(query)
        devices = cursor.fetchall()
        log_message(f"Retrieved {len(devices)} devices from database")
        return devices
    except Error as e:
        log_message(f"Error getting devices data: {e}")
        return []

def update_last_gps_assignment(connection, target_name, current_assignment, suggested_assignment):
    """
    Update last_gps_assignment in database when assignment changes
    Format: "Assignment Name MM/DD/YY"
    Returns the formatted string for display
    """
    try:
        # Get current date in MM/DD/YY format
        current_date = datetime.now().strftime('%m/%d/%y')
        
        # Format: "Assignment Name MM/DD/YY"
        last_assignment_entry = f"{current_assignment} {current_date}"
        
        # Update the database
        cursor = connection.cursor()
        update_query = """
        UPDATE devices 
        SET last_gps_assignment = %s 
        WHERE target_name = %s
        """
        cursor.execute(update_query, (last_assignment_entry, target_name))
        connection.commit()
        
        log_message(f"  Updated last_gps_assignment for {target_name}: {last_assignment_entry}")
        return last_assignment_entry
        
    except Error as e:
        log_message(f"  Error updating last_gps_assignment for {target_name}: {e}")
        return None

def extract_short_location(address):
    """Extract a short location from full address (last 2-3 components)"""
    if not address:
        return "Unknown Location"
    
    # Split by comma and get relevant parts
    parts = [p.strip() for p in address.split(',')]
    
    # Try to get province/region (usually last 2-3 parts)
    if len(parts) >= 2:
        # Get last 2 parts (typically: Province, Region)
        return ', '.join(parts[-2:])
    
    return address
    """Extract a short location from full address (last 2-3 components)"""
    if not address:
        return "Unknown Location"
    
    # Split by comma and get relevant parts
    parts = [p.strip() for p in address.split(',')]
    
    # Try to get province/region (usually last 2-3 parts)
    if len(parts) >= 2:
        # Get last 2 parts (typically: Province, Region)
        return ', '.join(parts[-2:])
    
    return address
    """Extract a short location from full address (last 2-3 components)"""
    if not address:
        return "Unknown Location"
    
    # Split by comma and get relevant parts
    parts = [p.strip() for p in address.split(',')]
    
    # Try to get province/region (usually last 2-3 parts)
    if len(parts) >= 2:
        # Get last 2 parts (typically: Province, Region)
        return ', '.join(parts[-2:])
    
    return address

def find_nearest_assignment(device, assignment_tables, connection):
    """
    Find the nearest assignment geofence for a device
    Returns assignment name if within 5km geofence, else returns "CURRENT LOC: short location"
    """
    device_location = (device['latitude'], device['longitude'])
    nearest_assignment = None
    nearest_distance = float('inf')
    nearest_site = None
    nearest_assignment_code = None
    
    # Assignment name mapping with codes
    assignment_map = {
        'assignment_amlan': {'name': 'Amlan', 'code': '0528'},
        'assignment_balingueo': {'name': 'Balingueo SS', 'code': '0000'},
        'assignment_banilad': {'name': 'Banilad SS', 'code': '0000'},
        'assignment_barotac': {'name': 'Barotac Viejo SS', 'code': '0000'},
        'assignment_bayombong': {'name': 'Bayombong SS', 'code': '0555'},
        'assignment_binan': {'name': 'Binan SS', 'code': '0000'},
        'assignment_bolo': {'name': 'Bolo', 'code': '0565'},
        'assignment_botolan': {'name': 'Botolan SS', 'code': '0569'},
        'assignment_cadiz': {'name': 'Cadiz SS', 'code': '0370'},
        'assignment_calacass': {'name': 'Calaca SS', 'code': '0313'},
        'assignment_calacatl': {'name': 'Calaca TL', 'code': '0311'},
        'assignment_calatrava': {'name': 'Calatrava SS', 'code': '0370'},
        'assignment_castillejos': {'name': 'Castillejos TL', 'code': '0557'},
        'assignment_dasmarinas': {'name': 'Dasmarinas SS', 'code': '0540'},
        'assignment_dumanjug': {'name': 'Dumanjug SS', 'code': '0352'},
        'assignment_ebmagalona': {'name': 'EB Magalona SS', 'code': '0370'},
        'assignment_headoffice': {'name': 'Head Office', 'code': 'HO'},
        'assignment_hermosatl': {'name': 'Hermosa TL', 'code': '0559'},
        'assignment_hermosa': {'name': 'Hermosa SS', 'code': '0555'},
        'assignment_ilijan': {'name': 'Ilijan SS', 'code': '0589'},
        'assignment_isabel': {'name': 'Isabel SS', 'code': '0511'},
        'assignment_maasin': {'name': 'Maasin SS', 'code': '0511'},
        'assignment_muntinlupa': {'name': 'Muntinlupa SS', 'code': '0000'},
        'assignment_pantabangan': {'name': 'Pantabangan SS', 'code': '0555'},
        'assignment_paoay': {'name': 'Paoay TL', 'code': 'PAOAY'},
        'assignment_pinamucan': {'name': 'Pinamucan SS', 'code': '0546'},
        'assignment_quirino': {'name': 'Quirino', 'code': 'QUIRINO'},
        'assignment_sanjose': {'name': 'San Jose SS', 'code': '0512'},
        'assignment_tabango': {'name': 'Tabango SS', 'code': '0511'},
        'assignment_tayabas': {'name': 'Tayabas SS', 'code': '0569'},
        'assignment_taytay': {'name': 'Taytay SS', 'code': '0555'},
        'assignment_terrasolar': {'name': 'Terra Solar', 'code': 'TERRA SOLAR'},
        'assignment_tuguegarao': {'name': 'Tuguegarao SS', 'code': '0555'},
        'assignment_tuy': {'name': 'Tuy SS', 'code': '0313'}
    }
    
    for table_name in assignment_tables:
        coordinates_data = get_assignment_coordinates(connection, table_name)
        
        for coord_info in coordinates_data:
            geofence_location = (coord_info['lat'], coord_info['lon'])
            distance = haversine(device_location, geofence_location)
            within_fence = is_within_geofence(device_location, geofence_location, 5)  # 5km radius
            
            if within_fence and distance < nearest_distance:
                nearest_distance = distance
                assignment_info = assignment_map.get(table_name, {'name': table_name.replace('assignment_', '').title(), 'code': ''})
                nearest_assignment = assignment_info['name']
                nearest_assignment_code = assignment_info['code']
                nearest_site = coord_info['site']
    
    if nearest_assignment:
        # Format with code if available
        if nearest_assignment_code:
            formatted_assignment = f"{nearest_assignment_code} - {nearest_assignment}"
        else:
            formatted_assignment = nearest_assignment
            
        return {
            'assignment': formatted_assignment,
            'distance': round(nearest_distance, 2),
            'site': nearest_site
        }
    else:
        # Return "CURRENT LOC: short location" if no assignment found
        short_location = extract_short_location(device.get('address', ''))
        return {
            'assignment': f"CURRENT LOC: {short_location}",
            'distance': None,
            'site': None
        }

def export_to_excel(devices_with_assignments, output_file):
    """Export data to Excel file grouped by Equipment Type and sorted by Assignment"""
    try:
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Create DataFrame
        df = pd.DataFrame(devices_with_assignments)
        
        # Sort by Equipment Type first, then by Assignment
        df = df.sort_values(['equipment_type', 'suggested_assignment', 'target_name'])
        
        # Group by equipment type
        equipment_groups = df.groupby('equipment_type')
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            workbook = writer.book
            worksheet = workbook.create_sheet('Vehicle Info Report', 0)
            
            # Add title row
            current_datetime = datetime.now().strftime('%B %d, %Y at %I:%M %p')
            worksheet['A1'] = f'EQUIPMENTS / VEHICLES AS OF {current_datetime.upper()} based on GPS'
            worksheet.merge_cells('A1:G1')
            
            # Style the title
            title_cell = worksheet['A1']
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Define borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Define fill colors
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for headers
            black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')    # Black for separator
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')    # White for all data rows
            
            # Define font colors
            red_font = Font(color='FF0000')  # Red text for breakdown
            black_font = Font(color='000000')  # Black text for operational
            
            title_cell.border = thin_border
            title_cell.fill = yellow_fill
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 5   # TYPE
            worksheet.column_dimensions['B'].width = 40  # EQUIPMENT/VEHICLE
            worksheet.column_dimensions['C'].width = 20  # TAG
            worksheet.column_dimensions['D'].width = 40  # ASSIGNMENT
            worksheet.column_dimensions['E'].width = 15  # STATUS
            worksheet.column_dimensions['F'].width = 50  # REMARKS
            worksheet.column_dimensions['G'].width = 20  # SUMMARY
            
            current_row = 2
            
            # Process each equipment type group
            for equipment_type, group_df in equipment_groups:
                # Add black separator row before headers (except for first group)
                if current_row > 2:
                    for col_idx in range(1, 8):
                        separator_cell = worksheet.cell(row=current_row, column=col_idx)
                        separator_cell.fill = black_fill
                        separator_cell.border = thin_border
                    current_row += 1
                
                # Add section header row
                worksheet.cell(row=current_row, column=1).value = 'TYPE'
                worksheet.cell(row=current_row, column=2).value = 'EQUIPMENT/VEHICLE'
                worksheet.cell(row=current_row, column=3).value = 'TAG'
                worksheet.cell(row=current_row, column=4).value = 'ASSIGNMENT'
                worksheet.cell(row=current_row, column=5).value = 'STATUS'
                worksheet.cell(row=current_row, column=6).value = 'REMARKS'
                worksheet.cell(row=current_row, column=7).value = 'SUMMARY'
                
                # Style header row with yellow background
                header_font = Font(bold=True)
                for col_idx in range(1, 8):
                    cell = worksheet.cell(row=current_row, column=col_idx)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.fill = yellow_fill
                
                current_row += 1
                group_start_row = current_row
                
                # Count statistics for this equipment type
                total_count = 0
                operational_count = 0
                breakdown_count = 0
                no_gps_count = 0
                assignment_counts = {}
                
                # Add data rows for this equipment type
                for _, row in group_df.iterrows():
                    # Use suggested_assignment as the assignment
                    new_assignment = row['suggested_assignment']
                    
                    # Check if we need to add "CURRENT LOC:" prefix using cut_address
                    if new_assignment.startswith("CURRENT LOC:"):
                        new_assignment = f"CURRENT LOC: {row['cut_address']}"
                    
                    # Handle empty values
                    tag_value = row.get('tag', '')
                    if tag_value is None or str(tag_value).strip() == '' or str(tag_value).lower() == 'none':
                        tag_value = ''
                    
                    status_value = row.get('physical_status', '')
                    if status_value is None or str(status_value).strip() == '' or str(status_value).lower() == 'none':
                        status_value = ''
                    
                    # Check days_no_gps and append "(No GPS)" to status
                    days_no_gps = row.get('days_no_gps', 0)
                    try:
                        days_no_gps = int(days_no_gps) if days_no_gps is not None else 0
                    except (ValueError, TypeError):
                        days_no_gps = 0
                    
                    if days_no_gps >= 60:
                        no_gps_count += 1
                        if status_value.strip():
                            status_value = f"{status_value} (No GPS)"
                        else:
                            status_value = "(No GPS)"
                    
                    remarks_value = row.get('remarks', '')
                    if remarks_value is None or str(remarks_value).strip() == '' or str(remarks_value).lower() == 'none':
                        remarks_value = ''
                    
                    # Check if we should display last_gps_assignment
                    last_gps_assignment = row.get('last_gps_assignment', '')
                    if last_gps_assignment and str(last_gps_assignment).strip() and str(last_gps_assignment).lower() != 'none':
                        last_assignment_text = f"Last Assignment: {last_gps_assignment}"
                        if remarks_value.strip():
                            remarks_value = f"{remarks_value}, {last_assignment_text}"
                        else:
                            remarks_value = last_assignment_text
                    
                    # Count statistics
                    total_count += 1
                    original_status = row.get('physical_status', '').upper()
                    if 'OPERATIONAL' in original_status or status_value.upper().startswith('OPERATIONAL'):
                        operational_count += 1
                    elif 'BREAKDOWN' in original_status or 'BREAKDOWN' in status_value.upper():
                        breakdown_count += 1
                    
                    # Count assignments (only if not CURRENT LOC)
                    if not new_assignment.startswith('CURRENT LOC:'):
                        if new_assignment in assignment_counts:
                            assignment_counts[new_assignment] += 1
                        else:
                            assignment_counts[new_assignment] = 1
                    
                    # Write row data
                    worksheet.cell(row=current_row, column=1).value = equipment_type
                    worksheet.cell(row=current_row, column=2).value = row['target_name']
                    worksheet.cell(row=current_row, column=3).value = tag_value
                    worksheet.cell(row=current_row, column=4).value = new_assignment
                    worksheet.cell(row=current_row, column=5).value = status_value
                    worksheet.cell(row=current_row, column=6).value = remarks_value
                    
                    # Determine if this row is breakdown (for red text)
                    is_breakdown = 'BREAKDOWN' in status_value.upper()
                    row_font = red_font if is_breakdown else black_font
                    
                    # Apply borders, alignment, colors, and font
                    for col_idx in range(1, 8):
                        cell = worksheet.cell(row=current_row, column=col_idx)
                        cell.border = thin_border
                        cell.fill = white_fill  # All data rows are white
                        cell.font = row_font    # Red text for breakdown, black for operational
                        
                        # Different alignment for different columns
                        if col_idx == 1:  # TYPE column - vertical text
                            cell.alignment = Alignment(vertical='center', text_rotation=90)
                        elif col_idx in [4, 5]:  # ASSIGNMENT and STATUS - no wrap
                            cell.alignment = Alignment(vertical='center', wrap_text=False)
                        else:  # Other columns - wrap text
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    current_row += 1
                
                # Merge TYPE column for this equipment group
                if current_row > group_start_row:
                    worksheet.merge_cells(start_row=group_start_row, start_column=1,
                                        end_row=current_row-1, end_column=1)
                    type_cell = worksheet.cell(row=group_start_row, column=1)
                    type_cell.alignment = Alignment(
                        horizontal='center', vertical='center', text_rotation=90
                    )
                
                # Add summary row
                summary_text = f"TOTAL: {total_count} OPERATIONAL: {operational_count} BREAKDOWN: {breakdown_count}"
                if no_gps_count > 0:
                    summary_text += f" NO GPS: {no_gps_count}"
                
                worksheet.cell(row=current_row, column=1).value = summary_text
                worksheet.merge_cells(start_row=current_row, start_column=1,
                                    end_row=current_row, end_column=6)
                cell = worksheet.cell(row=current_row, column=1)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = yellow_fill
                for col_idx in range(1, 7):
                    summary_cell = worksheet.cell(row=current_row, column=col_idx)
                    summary_cell.border = thin_border
                    summary_cell.fill = yellow_fill
                
                current_row += 1
                
                # Add assignment summary on the right side with merged cells
                summary_row_start = group_start_row
                for assignment, count in sorted(assignment_counts.items()):
                    # Merge cells based on count (e.g., if count is 3, merge 3 cells)
                    if count > 1:
                        worksheet.merge_cells(start_row=summary_row_start, start_column=7,
                                            end_row=summary_row_start + count - 1, end_column=7)
                    
                    # Set the value and style for the merged cell
                    summary_cell = worksheet.cell(row=summary_row_start, column=7)
                    summary_cell.value = f"{assignment} - {count}"
                    summary_cell.alignment = Alignment(horizontal='left', vertical='center')
                    summary_cell.border = thin_border
                    summary_cell.fill = white_fill
                    
                    # Apply borders to all cells in the merged range
                    for row_offset in range(count):
                        cell = worksheet.cell(row=summary_row_start + row_offset, column=7)
                        cell.border = thin_border
                        cell.fill = white_fill
                    
                    summary_row_start += count
                
                # Add blank row between equipment types (removed - black separator handles this)
                # current_row += 1
            
            # Freeze panes at row 2
            worksheet.freeze_panes = 'A2'
        
        log_message(f"Successfully exported to {output_file}")
        return True
    except Exception as e:
        log_message(f"Error exporting to Excel: {e}")
        return False

def main():
    """Main function"""
    log_message("=" * 80)
    log_message("Starting Vehicle Info Report Generator")
    log_message("=" * 80)
    
    # Connect to database
    connection = create_connection()
    if not connection:
        log_message("Failed to connect to database. Exiting.")
        print(json.dumps({'success': False, 'error': 'Database connection failed'}))
        sys.exit(1)
    
    try:
        # Get assignment tables
        log_message("Step 1: Retrieving assignment tables...")
        assignment_tables = get_assignment_tables(connection)
        if not assignment_tables:
            log_message("No assignment tables found. Exiting.")
            print(json.dumps({'success': False, 'error': 'No assignment tables found'}))
            return
        
        # Get devices data
        log_message("Step 2: Retrieving devices data...")
        devices = get_devices_data(connection)
        if not devices:
            log_message("No devices found. Exiting.")
            print(json.dumps({'success': False, 'error': 'No devices found'}))
            return
        
        # Process each device
        log_message("Step 3: Processing devices and finding assignments...")
        devices_with_assignments = []
        
        # Assignment name mapping for current assignments
        assignment_map = {
            'assignment_amlan': {'name': 'Amlan', 'code': '0528'},
            'assignment_balingueo': {'name': 'Balingueo SS', 'code': '0000'},
            'assignment_banilad': {'name': 'Banilad SS', 'code': '0000'},
            'assignment_barotac': {'name': 'Barotac Viejo SS', 'code': '0000'},
            'assignment_bayombong': {'name': 'Bayombong SS', 'code': '0555'},
            'assignment_binan': {'name': 'Binan SS', 'code': '0000'},
            'assignment_bolo': {'name': 'Bolo', 'code': '0565'},
            'assignment_botolan': {'name': 'Botolan SS', 'code': '0569'},
            'assignment_cadiz': {'name': 'Cadiz SS', 'code': '0370'},
            'assignment_calacass': {'name': 'Calaca SS', 'code': '0313'},
            'assignment_calacatl': {'name': 'Calaca TL', 'code': '0311'},
            'assignment_calatrava': {'name': 'Calatrava SS', 'code': '0370'},
            'assignment_castillejos': {'name': 'Castillejos TL', 'code': '0557'},
            'assignment_dasmarinas': {'name': 'Dasmarinas SS', 'code': '0540'},
            'assignment_dumanjug': {'name': 'Dumanjug SS', 'code': '0352'},
            'assignment_ebmagalona': {'name': 'EB Magalona SS', 'code': '0370'},
            'assignment_headoffice': {'name': 'Head Office', 'code': 'HO'},
            'assignment_hermosatl': {'name': 'Hermosa TL', 'code': '0559'},
            'assignment_hermosa': {'name': 'Hermosa SS', 'code': '0555'},
            'assignment_ilijan': {'name': 'Ilijan SS', 'code': '0589'},
            'assignment_isabel': {'name': 'Isabel SS', 'code': '0511'},
            'assignment_maasin': {'name': 'Maasin SS', 'code': '0511'},
            'assignment_muntinlupa': {'name': 'Muntinlupa SS', 'code': '0000'},
            'assignment_pantabangan': {'name': 'Pantabangan SS', 'code': '0555'},
            'assignment_paoay': {'name': 'Paoay TL', 'code': 'PAOAY'},
            'assignment_pinamucan': {'name': 'Pinamucan SS', 'code': '0546'},
            'assignment_quirino': {'name': 'Quirino', 'code': 'QUIRINO'},
            'assignment_sanjose': {'name': 'San Jose SS', 'code': '0512'},
            'assignment_tabango': {'name': 'Tabango SS', 'code': '0511'},
            'assignment_tayabas': {'name': 'Tayabas SS', 'code': '0569'},
            'assignment_taytay': {'name': 'Taytay SS', 'code': '0555'},
            'assignment_terrasolar': {'name': 'Terra Solar', 'code': 'TERRA SOLAR'},
            'assignment_tuguegarao': {'name': 'Tuguegarao SS', 'code': '0555'},
            'assignment_tuy': {'name': 'Tuy SS', 'code': '0313'}
        }
        
        for idx, device in enumerate(devices, 1):
            log_message(f"Processing {idx}/{len(devices)}: {device['target_name']}")
            
            # Skip GPS Trackers
            if device['equipment_type'] == "GPS Tracker":
                log_message(f"  Skipping GPS Tracker: {device['target_name']}")
                continue
            
            # Find nearest assignment
            assignment_info = find_nearest_assignment(
                device, assignment_tables, connection
            )
            
            # Map current assignment to friendly name with code
            current_assignment = device.get('current_assignment', '')
            assignment_data = assignment_map.get(current_assignment, {'name': current_assignment if current_assignment else 'N/A', 'code': ''})
            
            # Format current assignment with code
            if assignment_data['code']:
                current_assignment_display = f"{assignment_data['code']} - {assignment_data['name']}"
            else:
                current_assignment_display = assignment_data['name']
            
            suggested_assignment = assignment_info['assignment']
            
            # Check if assignment has changed
            assignment_changed = False
            last_gps_assignment = device.get('last_gps_assignment', '')
            
            # Parse last_gps_assignment to get the stored assignment name (without date)
            stored_last_assignment = ''
            if last_gps_assignment and str(last_gps_assignment).strip():
                # Format is "Assignment Name MM/DD/YY", extract just the assignment name
                parts = str(last_gps_assignment).rsplit(' ', 1)  # Split from the right to get date
                if len(parts) > 0:
                    stored_last_assignment = parts[0].strip()
            
            # Assignment has changed if:
            # 1. Suggested assignment is different from current assignment (and suggested is not CURRENT LOC)
            # 2. OR suggested assignment starts with "CURRENT LOC:" (assignment changed to unknown location)
            # 3. AND the stored last assignment is different from the current assignment (to update date)
            if current_assignment and suggested_assignment:
                # Normalize current assignment for comparison (use the full formatted version)
                current_normalized = current_assignment_display
                
                if suggested_assignment.startswith('CURRENT LOC:'):
                    # Assignment changed to current location (outside any geofence)
                    # Only update if we haven't recorded this change yet OR if stored assignment is different
                    if stored_last_assignment != current_normalized:
                        assignment_changed = True
                        log_message(f"  Assignment changed: {current_normalized} -> CURRENT LOC")
                elif current_normalized != suggested_assignment:
                    # Assignment changed to a different geofence
                    # Only update if we haven't recorded this change yet OR if stored assignment is different
                    if stored_last_assignment != current_normalized:
                        assignment_changed = True
                        log_message(f"  Assignment changed: {current_normalized} -> {suggested_assignment}")
            
            # Update last_gps_assignment if assignment has changed
            if assignment_changed and current_assignment:
                updated_last_assignment = update_last_gps_assignment(connection, device['target_name'], current_assignment_display, suggested_assignment)
                if updated_last_assignment:
                    last_gps_assignment = updated_last_assignment
            
            devices_with_assignments.append({
                'target_name': device['target_name'],
                'equipment_type': device['equipment_type'],
                'current_assignment': current_assignment_display,
                'suggested_assignment': suggested_assignment,
                'distance_km': assignment_info['distance'],
                'nearest_site': assignment_info['site'] if assignment_info['site'] else '',
                'address': device.get('address', ''),
                'cut_address': device.get('cut_address', ''),
                'tag': device.get('tag'),
                'physical_status': device.get('physical_status'),
                'remarks': device.get('remarks'),
                'days_no_gps': device.get('days_no_gps', 0),
                'last_gps_assignment': last_gps_assignment
            })
            
            log_message(f"  Suggested Assignment: {suggested_assignment}")
            if assignment_info['distance']:
                log_message(f"  Distance: {assignment_info['distance']} km")
            
            # Log GPS status if no GPS
            days_no_gps = device.get('days_no_gps', 0)
            if days_no_gps and days_no_gps >= 60:
                log_message(f"  GPS Status: No GPS ({days_no_gps} days)")
        
        # Export to Excel
        log_message(f"Step 4: Exporting {len(devices_with_assignments)} records to Excel...")
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(current_directory, f'vehicle_info_report_{timestamp}.xlsx')
        
        if export_to_excel(devices_with_assignments, output_file):
            log_message(f"SUCCESS: Vehicle info report exported to {output_file}")
            log_message(f"Total records processed: {len(devices_with_assignments)}")
            # Output JSON response only
            print(json.dumps({
                'success': True,
                'file': output_file,
                'records': len(devices_with_assignments)
            }))
        else:
            log_message("Failed to export to Excel")
            print(json.dumps({'success': False, 'error': 'Failed to export to Excel'}))
            sys.exit(1)
    
    except Exception as e:
        log_message(f"Error in main process: {e}")
        print(json.dumps({'success': False, 'error': str(e)}))
        sys.exit(1)
    
    finally:
        if connection and connection.is_connected():
            connection.close()
            log_message("Database connection closed")
        log_message("Script completed")

if __name__ == "__main__":
    main()